!pip install bcrypt

import sqlite3
import gradio as gr
import pandas as pd
import json, os, re
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
import openpyxl
import bcrypt


# ====== Database ======
def init_db():
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash BLOB NOT NULL,
            role TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

def add_user(username, password, role):
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt())
    try:
        c.execute("INSERT INTO users VALUES (?, ?, ?)", (username, hashed, role))
        conn.commit()
    except sqlite3.IntegrityError:
        pass
    conn.close()

def check_login(username, password):
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("SELECT password_hash, role FROM users WHERE username = ?", (username,))
    result = c.fetchone()
    conn.close()
    if result:
        stored_hash = result[0] if isinstance(result[0], bytes) else result[0].encode()
        if bcrypt.checkpw(password.encode(), stored_hash):
            return True, result[1]
    return False, None

def logout_action():
    return (
        gr.update(visible=True),   # halaman_login
        gr.update(visible=False),  # halaman_pilihan
        gr.update(visible=False),  # halaman_debitur
        gr.update(visible=False),  # halaman_karyawan
        "",  # username_in
        ""   # password_in
    )



replacement_nama_fasilitas = {
    "AEON Credit Services Indonesia": "AEON Credit",
    "Adira Dinamika Multi Finance": "Adira",
    "Akulaku Finance Indonesia": "Akulaku",
    "Atome Finance Indonesia": "Atome Finance",
    "Astra Multi Finance": "Astra MF",
    "BFI Finance Indonesia": "BFI",
    "BIMA Multi Finance": "Bima MF",
    "BPD Jawa Barat dan Banten": "BJB",
    "BPD Jawa Barat dan Banten Syariah": "BJB Syariah",
    "BPD Jawa Timur": "Bank Jatim",
    "BPD Sumatera Utara": "Bank Sumut",
    "Bank BCA Syariah": "BCA Syariah",
    "Bank CIMB Niaga": "CIMB Niaga",
    "Bank Central Asia": "BCA",
    "Bank DBS Indonesia": "Bank DBS",
    "Bank Danamon Indonesia": "Danamon",
    "Bank Danamon Indonesia Syariah": "Danamon Syariah",
    "Bank Hibank Indonesia": "Hibank",
    "Bank HSBC Indonesia": "HSBC",
    "Bank KEB Hana Indonesia": "Bank KEB Hana",
    "Bank Mandiri": "Bank Mandiri",
    "Bank Mandiri Taspen": "Bank Mantap",
    "Bank Mayapada Internasional": "Bank Mayapada",
    "Bank Maybank Indonesia": "Maybank",
    "Bank Mega Syariah": "Bank Mega Syariah",
    "Bank Muamalat Indonesia": "Bank Muamalat",
    "Bank Negara Indonesia": "BNI",
    "Bank Neo Commerce": "Akulaku",
    "Bank OCBC NISP": "OCBC NISP",
    "Bank Panin Indonesia": "Panin Bank",
    "Bank Permata": "Bank Permata",
    "Bank QNB Indonesia": "Bank QNB",
    "Bank Rakyat Indonesia": "BRI",
    "Bank Sahabat Sampoerna": "Bank Sampoerna",
    "Bank Saqu Indonesia (": "Bank Saqu",
    "Bank Seabank Indonesia": "Seabank",
    "Bank SMBC Indonesia": "Bank SMBC",
    "Bank Syariah Indonesia": "BSI",
    "Bank Tabungan Negara": "BTN",
    "Bank UOB Indonesia": "Bank UOB",
    "Bank Woori Saudara": "BWS",
    "Bank Woori Saudara Indonesia 1906":"BWS",
    "Bussan Auto Finance": "BAF",
    "Cakrawala Citra Mega Multifinance":"CCM Finance",
    "Commerce Finance": "Seabank",
    "Dana Mandiri Sejahtera": "Dana Mandiri",
    "Esta Dana Ventura": "Esta Dana",
    "Federal International Finance": "FIF",
    "Globalindo Multi Finance": "Globalindo MF",
    "Home Credit Indonesia": "Home Credit",
    "Indodana Multi Finance": "Indodana MF",
    "Indomobil Finance Indonesia": "IMFI",
    "Indonesia Airawata Finance (": "Indonesia Airawata Finance",
    "JACCS Mitra Pinasthika Mustika Finance Indonesia": "JACCS",
    "KB Finansia Multi Finance": "Kreditplus",
    "Kredivo Finance Indonesia": "Kredivo",
    "Krom Bank Indonesia": "Krom Bank",
    "LOLC Ventura Indonesia": "LOVI",
    "Mandala Multifinance": "Mandala MF",
    "Mandiri Utama Finance": "MUF",
    "Maybank Syariah": "Maybank Syariah",
    "Mega Auto Finance": "MAF",
    "Mega Central Finance": "MCF",
    "Mitra Bisnis Keluarga Ventura": "MBK",
    "Multifinance Anak Bangsa": "MF Anak Bangsa",
    "Panin Bank": "Panin Bank",
    "Permodalan Nasional Madani": "PNM",
    "Pratama Interdana Finance": "Pratama Finance",
    "Standard Chartered Bank": "Standard Chartered",
    "Summit Oto Finance": "Summit Oto",
    "Super Bank Indonesia": "Superbank",
    "Wahana Ottomitra Multiartha": "WOM",
    "Bank Jago": "Bank Jago",
    "Bank BTPN Syariah,": "BTPNS",
    "Bina Artha Ventura": "BAV"
}


def bersihkan_nama_fasilitas(nama_fasilitas: str) -> str:
    if not nama_fasilitas:
        return ""
    lower_fasilitas = nama_fasilitas.lower()
    if "d/h" in lower_fasilitas:
        nama_bersih = nama_fasilitas[:lower_fasilitas.find("d/h")].strip()
    elif "d.h" in lower_fasilitas:
        nama_bersih = nama_fasilitas[:lower_fasilitas.find("d.h")].strip()
    else:
        nama_bersih = nama_fasilitas.strip()
    for pattern in ["PT ", "PT.", "PD.", "(Persero)", "(Perseroda)", "Perseroda", "(UUS)", " Tbk"]:
        nama_bersih = nama_bersih.replace(pattern, "")
    nama_bersih = nama_bersih.replace("Bank Perekonomian Rakyat Syariah", "BPRS")
    nama_bersih = nama_bersih.replace("Bank Perekonomian Rakyat", "BPR")
    nama_bersih = nama_bersih.replace("Koperasi Simpan Pinjam", "KSP")
    nama_bersih = nama_bersih.strip()
    for nama_asli, alias in replacement_nama_fasilitas.items():
        if nama_asli.lower() == nama_bersih.lower():
            return alias
    return nama_bersih


def gabungkan_fasilitas_dengan_jumlah(fasilitas_list):
    counter = Counter(fasilitas_list)
    return '; '.join([f"{nama} ({jumlah})" if jumlah > 1 else nama for nama, jumlah in counter.items()])



def proses_files_debitur(files):
    if not files:
        return pd.DataFrame(), None

    hasil_semua = []
    excluded_fasilitas = {"BTPNS", "Bank Jago", "BAV"}

    for f in files:
        # === Identifikasi nama dan path file ===
        original_name = (
            getattr(f, "orig_name", None)
            or getattr(f, "name", None)
            or os.path.basename(f.name if hasattr(f, "name") else f)
        )
        path = getattr(f, "name", None) or getattr(f, "path", None) or f

        # Hanya proses file .txt
        if not str(original_name).lower().endswith(".txt"):
            continue

        # === Baca file JSON ===
        try:
            with open(path, "r", encoding="latin-1") as file:
                data = json.load(file)
        except Exception as e:
            print(f"Gagal membaca file: {original_name} -> {e}")
            continue

        # === Ambil data pokok debitur ===
        fasilitas = data.get("individual", {}).get("fasilitas", {}).get("kreditPembiayan", [])
        data_pokok = data.get("individual", {}).get("dataPokokDebitur", [])
        nama_debitur = ", ".join(
            set(
                debitur.get("namaDebitur", "")
                for debitur in data_pokok
                if debitur.get("namaDebitur")
            )
        )

        # === Variabel agregasi ===
        total_plafon = 0
        total_baki_debet = 0
        jumlah_fasilitas_aktif = 0
        kol_1_list, kol_25_list, wo_list, lovi_list = [], [], [], []
        baki_debet_kol25wo = 0

        # === Proses setiap fasilitas ===
        for item in fasilitas:
            tahun_wo = ""
            kondisi_ket = (item.get("kondisiKet") or "").lower()
            nama_fasilitas = item.get("ljkKet") or ""
            nama_fasilitas_lower = nama_fasilitas.lower()

            baki_debet_val = int(item.get("bakiDebet", 0))
            tunggakan_pokok = int(item.get("tunggakanPokok", 0))
            tunggakan_bunga = int(item.get("tunggakanBunga", 0))
            denda_val = int(item.get("denda", 0))

            # Tentukan status aktif
            is_aktif = kondisi_ket in ["fasilitas aktif", "diblokir sementara"]
            if not is_aktif and kondisi_ket not in ["lunas", "dihapusbukukan", "hapus tagih"]:
                if any([baki_debet_val > 0, tunggakan_pokok > 0, tunggakan_bunga > 0, denda_val > 0]):
                    is_aktif = True

            # Skip lunas kecuali LOVI
            if kondisi_ket == "lunas" and "pt lolc ventura indonesia" not in nama_fasilitas_lower:
                continue

            jumlah_hari_tunggakan = int(item.get("jumlahHariTunggakan", 0))
            kualitas = item.get("kualitas", "")
            kol_value = f"{kualitas}/{jumlah_hari_tunggakan}" if jumlah_hari_tunggakan != 0 else kualitas
            tanggal_kondisi = item.get("tanggalKondisi", "")
            baki_debet = baki_debet_val

            # Hitung baki debet jika nol tapi masih ada tunggakan
            if kondisi_ket in ["dihapusbukukan", "hapus tagih"] or is_aktif:
                if baki_debet == 0:
                    if tunggakan_pokok > 0:
                        baki_debet = tunggakan_pokok
                    else:
                        baki_debet = tunggakan_bunga + denda_val
                    if baki_debet == 0:
                        kondisi_ket = "lunas"
                        is_aktif = False

            plafon_awal = int(item.get("plafonAwal", 0))
            nama_fasilitas_bersih = bersihkan_nama_fasilitas(nama_fasilitas)
            baki_debet_format = "{:,.0f}".format(baki_debet).replace(",", ".")

            # Tentukan teks fasilitas
            if is_aktif and kualitas == "1" and jumlah_hari_tunggakan <= 30:
                fasilitas_teks = nama_fasilitas_bersih
            elif is_aktif:
                fasilitas_teks = f"{nama_fasilitas_bersih} Kol {kol_value} {baki_debet_format}"
            elif kondisi_ket in ["dihapusbukukan", "hapus tagih"]:
                try:
                    tahun_wo = int(str(tanggal_kondisi)[:4])
                except:
                    tahun_wo = ""
                fasilitas_teks = f"{nama_fasilitas_bersih} WO {tahun_wo} {baki_debet_format}"
            else:
                fasilitas_teks = nama_fasilitas_bersih

            # Tentukan LOVI
            if kondisi_ket == "lunas":
                fasilitas_lovi = "Lunas"
            elif is_aktif:
                fasilitas_lovi = f"Kol {kol_value}"
            elif kondisi_ket in ["dihapusbukukan", "hapus tagih"]:
                fasilitas_lovi = f"WO {tahun_wo} {baki_debet_format}"
            else:
                fasilitas_lovi = nama_fasilitas_bersih

            # === Klasifikasi fasilitas ===
            if "pt lolc ventura indonesia" not in nama_fasilitas_lower:
                if is_aktif:
                    total_plafon += plafon_awal
                    total_baki_debet += baki_debet
                    jumlah_fasilitas_aktif += 1
                    if kualitas == "1" and jumlah_hari_tunggakan <= 30:
                        if jumlah_hari_tunggakan == 0:
                            kol_1_list.append(nama_fasilitas_bersih)
                        else:
                            kol_1_list.append(f"{nama_fasilitas_bersih} Kol {kualitas}/{jumlah_hari_tunggakan}")
                    else:
                        kol_25_list.append(fasilitas_teks)
                        if nama_fasilitas_bersih not in excluded_fasilitas:
                            baki_debet_kol25wo += baki_debet
                elif kondisi_ket in ["dihapusbukukan", "hapus tagih"]:
                    wo_list.append(fasilitas_teks)
                    if nama_fasilitas_bersih not in excluded_fasilitas:
                        baki_debet_kol25wo += baki_debet
            else:
                if is_aktif or kondisi_ket in ["lunas", "dihapusbukukan"]:
                    tanggal_akad_akhir = item.get("tanggalAkadAkhir", "")
                    if tanggal_akad_akhir:
                        if not lovi_list:
                            lovi_list.append({"keterangan": fasilitas_lovi, "tanggal": tanggal_akad_akhir})
                        elif tanggal_akad_akhir > lovi_list[0]["tanggal"]:
                            lovi_list[0] = {"keterangan": fasilitas_lovi, "tanggal": tanggal_akad_akhir}

        # === Rekomendasi awal ===
        if jumlah_fasilitas_aktif >= 0 and not kol_25_list and not wo_list and not lovi_list:
            rekomendasi = "OK"
        elif any("lunas" in lovi.get("keterangan", "").lower() or "kol 1" in lovi.get("keterangan", "").lower() for lovi in lovi_list):
            rekomendasi = "OK"
        elif jumlah_fasilitas_aktif >= 0 and baki_debet_kol25wo <= 250_000 and not lovi_list:
            rekomendasi = "OK"
        else:
            rekomendasi = "NOT OK"

        # Simpan hasil per file
        filename = os.path.basename(original_name or path)
        nik = os.path.splitext(filename)[0]
        if nik.upper().startswith("NIK_"):
            nik = nik[4:]

        hasil_semua.append({
            "NIK": "'" + nik,
            "Nama Debitur": nama_debitur,
            "Rekomendasi": rekomendasi,
            "Jumlah Fasilitas": jumlah_fasilitas_aktif,
            "Total Plafon Awal": total_plafon if jumlah_fasilitas_aktif > 0 else "",
            "Total Baki Debet": total_baki_debet if jumlah_fasilitas_aktif > 0 else "",
            "Kol 1": gabungkan_fasilitas_dengan_jumlah(kol_1_list),
            "Kol 2-5": "; ".join(kol_25_list),
            "WO/dihapusbukukan": "; ".join(wo_list),
            "LOVI": "; ".join([l.get("keterangan", "") for l in lovi_list]),
            "Baki Debet Kol25WO": baki_debet_kol25wo  # hanya untuk logika, nanti dihapus dari Excel
        })

    if not hasil_semua:
        return pd.DataFrame(), None

    # === Grouping berdasarkan NIK ===
    grouped = defaultdict(list)
    for row in hasil_semua:
        nik_key = row["NIK"][:17]
        grouped[nik_key].append(row)

    def gabung_kolom(key, is_numerik=False):
        if is_numerik:
            return sum(row[key] for row in rows if isinstance(row[key], (int, float)))
        gabungan = "; ".join(str(row[key]) for row in rows if row[key])
        return "; ".join(sorted(set(gabungan.split("; "))))

    hasil_digabung = []
    for nik_key, rows in grouped.items():
        if len(rows) == 1:
            hasil_digabung.append(rows[0])
        else:
            hasil_digabung.append({
                "NIK": nik_key,
                "Nama Debitur": gabung_kolom("Nama Debitur"),
                "Rekomendasi": gabung_kolom("Rekomendasi"),
                "Jumlah Fasilitas": gabung_kolom("Jumlah Fasilitas", is_numerik=True),
                "Total Plafon Awal": gabung_kolom("Total Plafon Awal", is_numerik=True),
                "Total Baki Debet": gabung_kolom("Total Baki Debet", is_numerik=True),
                "Kol 1": gabung_kolom("Kol 1"),
                "Kol 2-5": gabung_kolom("Kol 2-5"),
                "WO/dihapusbukukan": gabung_kolom("WO/dihapusbukukan"),
                "LOVI": gabung_kolom("LOVI"),
                "Baki Debet Kol25WO": gabung_kolom("Baki Debet Kol25WO", is_numerik=True)
            })

    # === Pengecekan rekomendasi ulang ===
    hasil_final = []
    for row in hasil_digabung:
        jumlah_fasilitas_aktif = row["Jumlah Fasilitas"]
        baki_debet_kol25wo = row.get("Baki Debet Kol25WO", 0)

        kol_25_list = row["Kol 2-5"].split("; ") if row["Kol 2-5"] else []
        wo_list = row["WO/dihapusbukukan"].split("; ") if row["WO/dihapusbukukan"] else []
        lovi_list = row["LOVI"].split("; ") if row["LOVI"] else []

        if jumlah_fasilitas_aktif >= 0 and not kol_25_list and not wo_list and not lovi_list:
            rekomendasi = "OK"
        elif any("lunas" in lovi.lower() or "kol 1" in lovi.lower() for lovi in lovi_list):
            rekomendasi = "OK"
        elif jumlah_fasilitas_aktif >= 0 and baki_debet_kol25wo <= 250_000 and not lovi_list:
            rekomendasi = "OK"
        else:
            rekomendasi = "NOT OK"

        row["Rekomendasi"] = rekomendasi
        hasil_final.append(row)

    # === Output Excel ===
    df = pd.DataFrame(hasil_final)

    # Hapus kolom yang tidak perlu ditampilkan
    if "Baki Debet Kol25WO" in df.columns:
        df.drop(columns=["Baki Debet Kol25WO"], inplace=True)

    df.sort_values(by="NIK", inplace=True)
    tanggal_hari_ini = datetime.today().strftime("%d-%m-%Y_%H%M%S")
    output_file = f"Hasil SLIK Debitur {tanggal_hari_ini}.xlsx"
    df.to_excel(output_file, index=False)

    # === Format Excel ===
    custom_widths = {
        "NIK": 17,
        "Nama Debitur": 22,
        "Rekomendasi": 12,
        "Jumlah Fasilitas": 8,
        "Total Plafon Awal": 13,
        "Total Baki Debet": 13,
        "Kol 1": 31,
        "Kol 2-5": 31,
        "WO/dihapusbukukan": 31,
        "LOVI": 9
    }
    wrap_columns = set(custom_widths.keys())
    center_columns = set(custom_widths.keys())
    number_format_columns = {"Total Plafon Awal", "Total Baki Debet"}

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    header_row = ws[1]
    header = [cell.value for cell in header_row]

    for idx, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(idx)
        col_name = header[idx - 1] if idx - 1 < len(header) else ""

        wrap = col_name in wrap_columns
        center = col_name in center_columns

        if center and wrap:
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif center:
            alignment = Alignment(horizontal="center", vertical="center")
        elif wrap:
            alignment = Alignment(wrap_text=True)
        else:
            alignment = Alignment()

        for i, cell in enumerate(col_cells):
            cell.alignment = alignment
            cell.font = Font(size=8)
            cell.border = thin_border
            if i != 0 and col_name in number_format_columns:
                cell.number_format = "#,##0"

        if col_name in custom_widths:
            ws.column_dimensions[col_letter].width = custom_widths[col_name]
        else:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_file)

    return df, output_file

def proses_files_karyawan(files):
    if not files:
        return pd.DataFrame(), None

    hasil_semua = []

    for f in files:
        original_name = getattr(f, "orig_name", None) or getattr(f, "name", None) or os.path.basename(f.name if hasattr(f, "name") else f)
        path = getattr(f, "name", None) or getattr(f, "path", None) or f

        if not str(original_name).lower().endswith(".txt"):
            continue

        try:
            with open(path, "r", encoding="latin-1") as file:
                data = json.load(file)
        except Exception:
            continue

        fasilitas = data.get('individual', {}).get('fasilitas', {}).get('kreditPembiayan', [])
        data_pokok = data.get('individual', {}).get('dataPokokDebitur', [])
        nama_debitur = ', '.join(set(debitur.get('namaDebitur', '') for debitur in data_pokok if debitur.get('namaDebitur')))

        total_plafon = 0
        total_baki_debet = 0
        jumlah_fasilitas_aktif = 0
        kol_1_list, kol_2_list, kol_3_list, kol_4_list, kol_5_list, wo_list = [], [], [], [], [], []
        baki_debet_kol25wo = 0
        excluded_fasilitas = {"BTPNS", "Bank Jago", "BAV"}

        for item in fasilitas:
            kondisi_ket = (item.get('kondisiKet') or '').lower()
            nama_fasilitas = item.get('ljkKet') or ''
            nama_fasilitas_lower = nama_fasilitas.lower()

            baki_debet_val = int(item.get('bakiDebet', 0))
            tunggakan_pokok = int(item.get('tunggakanPokok', 0))
            tunggakan_bunga = int(item.get('tunggakanBunga', 0))
            denda_val = int(item.get('denda', 0))

            is_aktif = kondisi_ket in ['fasilitas aktif', 'diblokir sementara']

            if not is_aktif and kondisi_ket not in ['lunas', 'dihapusbukukan', 'hapus tagih']:
                if any([baki_debet_val > 0, tunggakan_pokok > 0, tunggakan_bunga > 0, denda_val > 0]):
                    is_aktif = True

            jumlah_hari_tunggakan = int(item.get('jumlahHariTunggakan', 0))
            kualitas = item.get('kualitas', '')
            kol_value = f"{kualitas}/{jumlah_hari_tunggakan}"
            tanggal_kondisi = item.get('tanggalKondisi', '')
            baki_debet = baki_debet_val

            if kondisi_ket in ['dihapusbukukan', 'hapus tagih'] or is_aktif:
                if baki_debet == 0:
                    if tunggakan_pokok > 0:
                        baki_debet = tunggakan_pokok
                    else:
                        baki_debet = tunggakan_bunga + denda_val
                    if baki_debet == 0:
                        kondisi_ket = 'lunas'
                        is_aktif = False

            plafon_awal = int(item.get('plafonAwal', 0))
            nama_fasilitas_bersih = bersihkan_nama_fasilitas(nama_fasilitas)
            baki_debet_format = "{:,.0f}".format(baki_debet).replace(",", ".")
            fasilitas_teks = f"{nama_fasilitas_bersih} Kol {kol_value} {baki_debet_format}"

            if is_aktif:
                if kualitas == '1':
                    if jumlah_hari_tunggakan <= 30:
                        kol_1_list.append(fasilitas_teks)
                    else:
                        kol_2_list.append(fasilitas_teks)
                elif kualitas == '2':
                    kol_2_list.append(fasilitas_teks)
                elif kualitas == '3':
                    kol_3_list.append(fasilitas_teks)
                elif kualitas == '4':
                    kol_4_list.append(fasilitas_teks)
                elif kualitas == '5':
                    kol_5_list.append(fasilitas_teks)

                if nama_fasilitas_bersih not in excluded_fasilitas:
                    baki_debet_kol25wo += baki_debet

                total_plafon += plafon_awal
                total_baki_debet += baki_debet
                jumlah_fasilitas_aktif += 1

            elif kondisi_ket in ['dihapusbukukan', 'hapus tagih']:
                wo_list.append(f"{nama_fasilitas_bersih} WO {tanggal_kondisi[:4]} {baki_debet_format}")
                if nama_fasilitas_bersih not in excluded_fasilitas:
                    baki_debet_kol25wo += baki_debet

        filename = os.path.basename(original_name or path)
        nik = os.path.splitext(filename)[0]
        if nik.upper().startswith("NIK_"):
            nik = nik[4:]

        hasil_semua.append({
            'NIK': "'" + nik,
            'Nama Karyawan': nama_debitur,
            'Jumlah Fasilitas': jumlah_fasilitas_aktif,
            'Total Plafon Awal': total_plafon if jumlah_fasilitas_aktif > 0 else "",
            'Total Baki Debet': total_baki_debet if jumlah_fasilitas_aktif > 0 else "",
            'Kol 1': gabungkan_fasilitas_dengan_jumlah(kol_1_list),
            'Kol 2': '; '.join(kol_2_list),
            'Kol 3': '; '.join(kol_3_list),
            'Kol 4': '; '.join(kol_4_list),
            'Kol 5': '; '.join(kol_5_list),
            'WO/dihapusbukukan': '; '.join(wo_list)
        })

    from collections import defaultdict
    grouped = defaultdict(list)
    for row in hasil_semua:
        nik_key = row['NIK'][:17]
        grouped[nik_key].append(row)

    hasil_digabung = []
    for nik_key, rows in grouped.items():
        if len(rows) == 1:
            hasil_digabung.append(rows[0])
            continue

        def gabung_kolom(key, is_numerik=False):
            if is_numerik:
                return sum(row[key] for row in rows if isinstance(row[key], (int, float)))
            gabungan = '; '.join(str(row[key]) for row in rows if row[key])
            return '; '.join(sorted(set(gabungan.split('; '))))

        hasil_digabung.append({
            'NIK': nik_key,
            'Nama Karyawan': gabung_kolom('Nama Karyawan'),
            'Jumlah Fasilitas': gabung_kolom('Jumlah Fasilitas', is_numerik=True),
            'Total Plafon Awal': gabung_kolom('Total Plafon Awal', is_numerik=True),
            'Total Baki Debet': gabung_kolom('Total Baki Debet', is_numerik=True),
            'Kol 1': gabung_kolom('Kol 1'),
            'Kol 2': gabung_kolom('Kol 2'),
            'Kol 3': gabung_kolom('Kol 3'),
            'Kol 4': gabung_kolom('Kol 4'),
            'Kol 5': gabung_kolom('Kol 5'),
            'WO/dihapusbukukan': gabung_kolom('WO/dihapusbukukan')
        })

    df = pd.DataFrame(hasil_digabung)
    df.sort_values(by='NIK', inplace=True)
    output_file = f'Hasil SLIK Karyawan {datetime.today().strftime("%d-%m-%Y_%H%M%S")}.xlsx'
    df.to_excel(output_file, index=False)

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    custom_widths = {
        'NIK': 17, 'Nama Karyawan': 22,  'Jumlah Fasilitas': 8,
        'Total Plafon Awal': 13, 'Total Baki Debet': 13, 'Kol 1': 31, 'Kol 2': 31, 'Kol 3': 31,
        'Kol 4': 31, 'Kol 5': 31, 'WO/dihapusbukukan': 31
    }
    wrap_columns = {'Nama Karyawan','Total Plafon Awal','Total Baki Debet','Kol 1', 'Kol 2', 'Kol 3', 'Kol 4', 'Kol 5', 'WO/dihapusbukukan'}
    center_columns = {'NIK','Nama Karyawan', 'Jumlah Fasilitas', 'Total Plafon Awal','Total Baki Debet','Kol 1', 'Kol 2', 'Kol 3', 'Kol 4', 'Kol 5', 'WO/dihapusbukukan'}
    number_format_columns = {'Total Plafon Awal', 'Total Baki Debet'}
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    header_row = ws[1]
    header = [cell.value for cell in header_row]
    for idx, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(idx)
        col_name = header[idx - 1] if idx - 1 < len(header) else ''
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=col_name in wrap_columns)
        for i, cell in enumerate(col_cells):
            cell.alignment = alignment
            cell.font = Font(size=8)
            cell.border = thin_border
            if i != 0 and col_name in number_format_columns:
                cell.number_format = '#,##0'
        if col_name in custom_widths:
            ws.column_dimensions[col_letter].width = custom_widths[col_name]

    wb.save(output_file)
    return df, output_file

def clear_data():
    return None, None, pd.DataFrame()







# ====== UI ======
with gr.Blocks(
    theme=gr.themes.Soft(),
    css="""
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@400;500;600&display=swap');

        * {
            font-family: 'Poppins', sans-serif !important;
        }

        h3 {
            text-align: center !important;
            margin-top: 60px !important;
            font-weight: normal !important;
            font-size: 12pt !important;
        }

        /* Navbar */
        .navbar {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 10px 20px;
            border-bottom: 1px solid #ddd;
            position: sticky;
            top: 0;
            z-index: 100;
        }

        .navbar-title {
            font-weight: 600;
            font-size: 20px;
        }

        .logout-btn {
            margin-left: auto;
        }

        /* Login Container */
        #login-container {
            max-width: 350px;
            margin: 80px auto;
            padding: 20px;
            border-radius: 10px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }

        #signin-title {
            text-align: center !important;
        }

        /* Button Row */
        .center-row {
            display: flex !important;
            justify-content: center !important;
            gap: 20px !important;
        }

        .image-button {
            width: 150px !important;
            height: 150px !important;
            min-width: 150px !important;
            max-width: 150px !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            padding: 0 !important;
            border-radius: 10px !important;
            font-size: 14px !important;
            text-align: center !important;
            white-space: normal !important;
            word-break: break-word !important;
        }

        .image-button:hover {
            transform: scale(1.05);
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }

        .small-text,
        .small-text * {
            font-size: 10px !important;
            color: #555 !important;
        }

        #preview-table,
        #preview-table * {
            font-size: 10px !important;
        }

        .tight-row {
            gap: 20px !important;
            justify-content: center;
            flex-wrap: wrap;
        }

        footer {
            text-align: center;
            margin-top: 10px;
            color: gray;
            font-size: 13px;
        }
    """
) as demo:


    # ====== Halaman Login ======
    with gr.Column(visible=True, elem_id="login-container") as halaman_login:
        gr.Markdown("SLIK Data Processing System", elem_id="signin-title")
        gr.Markdown("Sign In", elem_id="signin-title")
        username_in = gr.Textbox(label="Username")
        password_in = gr.Textbox(label="Password", type="password")
        login_btn = gr.Button("Login", variant="primary")
        login_msg = gr.Markdown("")
        gr.HTML("<footer>© 2025 | Created by Ayu Nurhasanah</footer>")


    # ====== Halaman Pilihan ======
    with gr.Column(visible=False) as halaman_pilihan:
        with gr.Row(elem_classes="navbar"):
            navbar_title = gr.Markdown("**SLIK Data Processing System**", elem_classes="navbar-title")
            logout_btn_pilihan = gr.Button("Logout", variant="secondary", size="sm", scale=0, min_width=100, elem_classes="logout-btn")

        gr.Markdown("<h3 style='margin-bottom: 30px;'>Select data type to continue</h3>")
        with gr.Row(elem_classes="center-row"):
            tombol_debitur = gr.Button("Debtor Data", elem_classes="image-button")
            tombol_karyawan = gr.Button("Employee Data", elem_classes="image-button")

        gr.HTML("<footer style='margin-top: 60px !important; '>© 2025 | Created by Ayu Nurhasanah</footer>")

    # ---------------- Halaman Debitur ----------------
    with gr.Column(visible=False) as halaman_debitur:
        with gr.Row(elem_classes="navbar"):
            navbar_title_debitur = gr.Markdown("**Debtor Data Processing**", elem_classes="navbar-title")
            logout_btn_debitur = gr.Button("Logout", variant="secondary", size="sm", scale=0, min_width=100, elem_classes="logout-btn")

        tombol_kembali_debitur = gr.Button("Back", size="sm", scale=0, min_width=100)
        gr.Markdown("Upload several .txt files, then click **Process**, then download the processed Excel file.")

        with gr.Row():
            with gr.Column(scale=1):
                inp_files_debitur = gr.File(label="", file_count="multiple", file_types=[".txt"])
                tombol_proses_debitur = gr.Button("Process", variant="primary")
            with gr.Column(scale=1):
                output_file_debitur = gr.File(label="Download", file_types=[".xlsx"])
                clear_btn_debitur = gr.Button("Clear Data", variant="secondary")

        output_df_debitur = gr.Dataframe(label="Preview", elem_id="preview-table", wrap=False)
        gr.HTML("<footer>© 2025 | Created by Ayu Nurhasanah</footer>")

    # ---------------- Halaman Karyawan ----------------
    with gr.Column(visible=False) as halaman_karyawan:
        with gr.Row(elem_classes="navbar"):
            navbar_title_karyawan = gr.Markdown("**Employee Data Processing**", elem_classes="navbar-title")
            logout_btn_karyawan = gr.Button("Logout", variant="secondary", size="sm", scale=0, min_width=100, elem_classes="logout-btn")

        tombol_kembali_karyawan = gr.Button("Back", size="sm", scale=0, min_width=100)
        gr.Markdown("Upload several .txt files, then click **Process**, then download the processed Excel file.")

        with gr.Row():
            with gr.Column(scale=1):
                inp_files_karyawan = gr.File(label="", file_count="multiple", file_types=[".txt"])
                tombol_proses_karyawan = gr.Button("Process", variant="primary")
            with gr.Column(scale=1):
                output_file_karyawan = gr.File(label="Download", file_types=[".xlsx"])
                clear_btn_karyawan = gr.Button("Clear Data", variant="secondary")

        output_df_karyawan = gr.Dataframe(label="Preview", elem_id="preview-table", wrap=False)
        gr.HTML("<footer>© 2025 | Created by Ayu Nurhasanah</footer>")

    # ====== Event Login ======
    def handle_login(username, password):
        valid, role = check_login(username, password)
        if valid:
            return (
                gr.update(visible=False),  # sembunyikan halaman login
                gr.update(visible=True),   # tampilkan halaman pilihan
                gr.update(visible=(role=="admin" or role=="user1")),  # tombol Debitur
                gr.update(visible=(role=="admin" or role=="user2")),  # tombol Karyawan
                "",  # reset username
                ""   # reset password
            )
        else:
            return (
                gr.update(visible=True),
                gr.update(visible=False),
                gr.update(visible=False),  # tombol Debitur
                gr.update(visible=False),  # tombol Karyawan
                username,
                ""
            )

    # Event Login
    login_btn.click(
        handle_login,
        inputs=[username_in, password_in],
        outputs=[
            halaman_login,
            halaman_pilihan,
            tombol_debitur,
            tombol_karyawan,
            username_in,
            password_in
        ]
    )

    # ====== Navigasi ======
    tombol_debitur.click(lambda: (gr.update(visible=False), gr.update(visible=True), gr.update(visible=False)),
                         outputs=[halaman_pilihan, halaman_debitur, halaman_karyawan])
    tombol_karyawan.click(lambda: (gr.update(visible=False), gr.update(visible=False), gr.update(visible=True)),
                          outputs=[halaman_pilihan, halaman_debitur, halaman_karyawan])
    tombol_kembali_debitur.click(lambda: (gr.update(visible=True), gr.update(visible=False), gr.update(visible=False)),
                                 outputs=[halaman_pilihan, halaman_debitur, halaman_karyawan])
    tombol_kembali_karyawan.click(lambda: (gr.update(visible=True), gr.update(visible=False), gr.update(visible=False)),
                                  outputs=[halaman_pilihan, halaman_debitur, halaman_karyawan])

    # Semua tombol logout kembali ke login
    for btn in [logout_btn_pilihan, logout_btn_debitur, logout_btn_karyawan]:
        btn.click(
            logout_action,
            outputs=[halaman_login, halaman_pilihan, halaman_debitur, halaman_karyawan, username_in, password_in]
        )
    # ====== Proses Data ======
    tombol_proses_debitur.click(fn=proses_files_debitur, inputs=[inp_files_debitur],
                                outputs=[output_df_debitur, output_file_debitur])
    clear_btn_debitur.click(fn=clear_data, outputs=[inp_files_debitur, output_file_debitur, output_df_debitur])

    tombol_proses_karyawan.click(fn=proses_files_karyawan, inputs=[inp_files_karyawan],
                                 outputs=[output_df_karyawan, output_file_karyawan])
    clear_btn_karyawan.click(fn=clear_data, outputs=[inp_files_karyawan, output_file_karyawan, output_df_karyawan])

# ====== Init Database & Default Users ======
init_db()
add_user("AYUNU", "Ayu@1234", "admin")
add_user("DICKYP", "Dicky@1234", "admin")
add_user("SITIN", "Siti@1234", "user1")
add_user("NINDAD", "Ninda@1234", "user1")
add_user("ZENAM", "Zena@1234", "user1")
add_user("YUDISTIRAR", "Yudistira@1234", "user1")
add_user("LUGIP", "Lugi@1234", "user1")
add_user("OKKYA", "Okky@1234", "user1")
add_user("SRIAYU", "Sri@1234", "user1")
add_user("ANNISAFA", "Annisa@1234", "user1")
add_user("SHOLEHHAA", "Sho@1234", "user2")

demo.launch(
    server_name="0.0.0.0",
    server_port=int(os.environ.get("PORT", 7860))
)
